"""Utilities for managing templates, export fields, and product generation."""
from __future__ import annotations

import csv
import json
import logging
import os
import re
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from app_paths import get_config_path
from errors import MissingDependencyError
from formula_engine import FormulaEngine, FormulaError

from database import collect_models, load_specs_map

try:
    from dateutil.relativedelta import relativedelta as _relativedelta
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    _relativedelta = None


class _CallableDateTime(datetime):
    """Datetime subclass that can be used both as value and callable."""

    def __new__(cls, value, *args, **kwargs):
        if isinstance(value, datetime) and not args and not kwargs:
            base = value
            extra = {}
            if hasattr(base, "fold"):
                extra["fold"] = getattr(base, "fold", 0)
            return datetime.__new__(
                cls,
                base.year,
                base.month,
                base.day,
                base.hour,
                base.minute,
                base.second,
                base.microsecond,
                tzinfo=base.tzinfo,
                **extra,
            )
        return datetime.__new__(cls, value, *args, **kwargs)

    def __call__(self):
        kwargs = {}
        if hasattr(self, "fold"):
            kwargs["fold"] = getattr(self, "fold", 0)
        return datetime(
            self.year,
            self.month,
            self.day,
            self.hour,
            self.minute,
            self.second,
            self.microsecond,
            tzinfo=self.tzinfo,
            **kwargs,
        )


def _relativedelta_helper(*args, **kwargs):
    if _relativedelta is None:
        raise RuntimeError(
            "Функція relativedelta недоступна. Встановіть пакет python-dateutil."
        )
    return _relativedelta(*args, **kwargs)

APP_TITLE = "Prom Generator"

DEPENDENCY_WARNINGS: List[str] = []

CSV_JSON_FALLBACK_NOTE = "\nЕкспорт у CSV (.csv) та JSON (.json) залишається доступним."
OPENPYXL_INSTALL_HINT = (
    "Встановіть бібліотеку командою 'pip install openpyxl' і перезапустіть застосунок, щоб увімкнути цей формат."
)

EXCEL_EXPORT_BLOCKED_MESSAGE = ""
OPENPYXL_IMPORT_ERROR_DETAIL = ""

try:
    from openpyxl import Workbook
except ModuleNotFoundError as exc:
    Workbook = None
    OPENPYXL_IMPORT_ERROR_DETAIL = str(exc)
    EXCEL_EXPORT_BLOCKED_MESSAGE = (
        "Експорт у формат Excel (.xlsx) недоступний: бібліотека openpyxl не встановлена."
    )
except ImportError as exc:  # pragma: no cover - defensive
    Workbook = None
    OPENPYXL_IMPORT_ERROR_DETAIL = str(exc)
    EXCEL_EXPORT_BLOCKED_MESSAGE = (
        "Експорт у формат Excel (.xlsx) недоступний: не вдалося завантажити бібліотеку openpyxl."
    )
else:
    OPENPYXL_IMPORT_ERROR_DETAIL = ""
    EXCEL_EXPORT_BLOCKED_MESSAGE = ""

OPENPYXL_AVAILABLE = Workbook is not None


class ExportError(RuntimeError):
    """Custom exception describing export failures."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message

    def __str__(self):  # pragma: no cover - mirrors base behaviour
        return self.message


EXPORT_ERR_NO_OPENPYXL = "NO_OPENPYXL"
EXPORT_ERR_FOLDER_PREP = "FOLDER_PREP_FAILED"
EXPORT_ERR_PERMISSION = "PERMISSION_DENIED"
EXPORT_ERR_OS_ERROR = "OS_ERROR"
EXPORT_ERR_UNKNOWN_FORMAT = "UNKNOWN_FORMAT"

if EXCEL_EXPORT_BLOCKED_MESSAGE:
    detail_suffix = (
        f"\nДеталі: {OPENPYXL_IMPORT_ERROR_DETAIL}" if OPENPYXL_IMPORT_ERROR_DETAIL else ""
    )
    DEPENDENCY_WARNINGS.append(
        EXCEL_EXPORT_BLOCKED_MESSAGE
        + detail_suffix
        + "\n"
        + OPENPYXL_INSTALL_HINT
        + CSV_JSON_FALLBACK_NOTE
    )
else:
    OPENPYXL_IMPORT_ERROR_DETAIL = ""

try:
    from jinja2 import Template, TemplateError
except ModuleNotFoundError as exc:
    raise MissingDependencyError(
        "Бібліотека Jinja2 не знайдена. Встановіть її командою 'pip install jinja2'."
    ) from exc

LOGGER = logging.getLogger(__name__)

TEMPLATES_FILENAME = "templates.json"
EXPORT_FIELDS_FILENAME = "export_fields.json"
TITLE_TAGS_FILENAME = "title_tags_templates.json"
FILM_TYPE_DEFAULT_LABEL = "Універсальний шаблон"
CATEGORY_SCOPE_DEFAULT_LABEL = "Для всіх категорій"
# Ключ для шаблонів опису, які застосовуються для всіх категорій
GLOBAL_DESCRIPTION_KEY = "__global__"
TEMPLATE_LANGUAGE_DEFAULT_LABEL = "За замовчуванням"

DEFAULT_TEMPLATE_LANGUAGES = [
    {"code": "uk", "label": "Українська"},
    {"code": "ru", "label": "Російська"},
    {"code": "en", "label": "English"},
]


def _config_path(filename: str) -> Path:
    path = get_config_path(filename)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _write_json_config(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _backup_corrupt_config(path: Path) -> Optional[Path]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}.bad_{ts}{path.suffix}")
    try:
        path.rename(backup)
        return backup
    except OSError:
        try:
            data = path.read_bytes()
        except OSError:
            LOGGER.exception("Не вдалося зчитати пошкоджений конфіг %s для бекапу", path)
            return None
        try:
            backup.write_bytes(data)
            path.unlink(missing_ok=True)
            return backup
        except Exception:
            LOGGER.exception("Не вдалося створити бекап для %s", path)
            return None


def _load_json_config(path: Path, default_factory: Callable[[], object], expected_type) -> object:
    if not path.exists():
        data = default_factory()
        _write_json_config(path, data)
        return data
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as exc:
        backup = _backup_corrupt_config(path)
        LOGGER.warning(
            "Файл %s пошкоджено, створено бекап %s і відновлено значення за замовчуванням",
            path,
            backup,
            exc_info=exc,
        )
        data = default_factory()
        _write_json_config(path, data)
        return data
    if not isinstance(data, expected_type):
        data = default_factory()
        _write_json_config(path, data)
    return data


def _normalize_language_definitions(raw_languages):
    normalized = []
    seen = set()
    if isinstance(raw_languages, (list, tuple)):
        for item in raw_languages:
            code = None
            label = None
            if isinstance(item, dict):
                code = item.get("code") or item.get("id") or item.get("name")
                label = item.get("label") or item.get("title") or item.get("name")
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                code, label = item[0], item[1]
            elif isinstance(item, str):
                code = item
                label = item
            if not isinstance(code, str):
                continue
            code = code.strip()
            if not code or code in seen:
                continue
            if not isinstance(label, str):
                label = code
            label = label.strip()
            if not label:
                label = code
            normalized.append({"code": code, "label": label})
            seen.add(code)
    if not normalized:
        normalized = deepcopy(DEFAULT_TEMPLATE_LANGUAGES)
    return normalized


def _language_choices(languages):
    return [
        (item["code"], item.get("label", item["code"]))
        for item in languages
        if item.get("code")
    ]


def _normalize_template_language_entry(entry, fallback_value=""):
    if isinstance(entry, dict):
        default_value = entry.get("default")
        if not isinstance(default_value, str):
            default_value = fallback_value
        languages_block = entry.get("languages")
        normalized_languages = {}
        if isinstance(languages_block, dict):
            for code, value in languages_block.items():
                if isinstance(code, str) and isinstance(value, str):
                    normalized_languages[code] = value
        else:
            for code, value in entry.items():
                if code in {"default", "languages"}:
                    continue
                if isinstance(code, str) and isinstance(value, str):
                    normalized_languages[code] = value
        return {"default": default_value, "languages": normalized_languages}
    if isinstance(entry, str):
        return {"default": entry, "languages": {}}
    return {"default": fallback_value, "languages": {}}


def _get_language_template_value(entry, language_code, fallback_value=""):
    if isinstance(entry, dict):
        languages_block = entry.get("languages")
        if isinstance(languages_block, dict):
            value = languages_block.get(language_code)
            if isinstance(value, str):
                return value
        value = entry.get(language_code)
        if isinstance(value, str):
            return value
        default_value = entry.get("default")
        if isinstance(default_value, str):
            return default_value
    elif isinstance(entry, str):
        return entry
    return fallback_value


def _set_language_template_value(entry, language_code, value, fallback_value=""):
    normalized = _normalize_template_language_entry(entry, fallback_value=fallback_value)
    if isinstance(language_code, str) and language_code:
        normalized.setdefault("languages", {})[language_code] = value
    else:
        normalized["default"] = value
    return normalized


def _rename_language_in_entry(entry, old_code: str, new_code: Optional[str]) -> bool:
    if not isinstance(entry, dict) or not isinstance(old_code, str) or not old_code:
        return False
    changed = False
    languages_block = entry.get("languages")
    if isinstance(languages_block, dict) and old_code in languages_block:
        value = languages_block.pop(old_code)
        if isinstance(new_code, str) and new_code:
            languages_block[new_code] = value
        changed = True
    if old_code in entry and isinstance(entry.get(old_code), str):
        value = entry.pop(old_code)
        if isinstance(new_code, str) and new_code:
            entry[new_code] = value
        changed = True
    return changed


_LANGUAGE_SUFFIX_MAP = {
    "_укр": "uk",
    "_ua": "uk",
    "_uk": "uk",
    "_анг": "en",
    "_англ": "en",
    "_en": "en",
    "_eng": "en",
    "_рус": "ru",
    "_рос": "ru",
    "_ru": "ru",
}

_MISSING = object()


def _infer_export_language(field_name: str):
    if not isinstance(field_name, str):
        return None
    lowered = field_name.strip().lower()
    for suffix, code in _LANGUAGE_SUFFIX_MAP.items():
        if lowered.endswith(suffix):
            return code
    return None


def _normalize_export_field_languages(field_name: str, languages=_MISSING):
    infer_default = languages is _MISSING
    if languages is _MISSING:
        languages = None
    normalized: List[str] = []
    if isinstance(languages, str):
        languages = [languages]
    if isinstance(languages, (list, tuple, set)):
        seen = set()
        for lang in languages:
            if not isinstance(lang, str):
                continue
            code = lang.strip()
            if not code or code in seen:
                continue
            normalized.append(code)
            seen.add(code)
    if not normalized and infer_default:
        inferred = _infer_export_language(field_name)
        if inferred:
            normalized.append(inferred)
    return normalized


DEFAULT_TEMPLATES = {
    "title_template": "Гідрогелева плівка {{ film_type }} {{ brand }} {{ model }}",
    "tags_template": "{{ brand }} {{ model }}, плівка {{ brand }} {{ model }}, hydrogel film {{ brand }} {{ model }}, {{ film_type }} {{ brand }} {{ model }}",
    "descriptions": {
        GLOBAL_DESCRIPTION_KEY: {
            "default": "{{ brand }} {{ model }} — надійний захист екрана гідрогелевою плівкою."
        },
        "Смартфони": {
            "прозора": "Прозора плівка для {{ brand }} {{ model }} — базовий прозорий захист, висока чутливість та легка поклейка.",
            "матова": "Матова плівка для {{ brand }} {{ model }} — мінімум відблисків, комфорт на сонці, приємний тактильний ефект.",
            "anti-blue": "Anti-Blue плівка для {{ brand }} {{ model }} — фільтрація синього світла для зниження втоми очей.",
            "privacy clear": "Privacy Clear для {{ brand }} {{ model }} — захист від підглядання під прямим кутом, прозора фронтальна видимість.",
            "privacy mate": "Privacy Mate для {{ brand }} {{ model }} — матова з приватністю, менше відблисків і захист від бічних кутів огляду.",
            "default": "Універсальна плівка для {{ brand }} {{ model }} — захист від подряпин та відбитків."
        },
        "Планшети": {
            "прозора": "Прозора плівка для планшета {{ brand }} {{ model }} — чиста картинка на великому екрані, прост поклейка.",
            "матова": "Матова плівка для планшета {{ brand }} {{ model }} — мінімум відблисків, комфорт для роботи/навчання.",
            "anti-blue": "Anti-Blue для планшета {{ brand }} {{ model }} — зниження синього спектру, довша робота без втоми очей.",
            "default": "Універсальна плівка для планшета {{ brand }} {{ model }} — збалансований захист поверхні."
        }
    },
    "film_types": [
        {"name": "прозора", "enabled": True},
        {"name": "матова", "enabled": True},
        {"name": "privacy clear", "enabled": True},
        {"name": "privacy mate", "enabled": True},
        {"name": "anti-blue", "enabled": True}
    ],
    "template_languages": deepcopy(DEFAULT_TEMPLATE_LANGUAGES),
}

DEFAULT_EXPORT_FIELDS = [
    {"field": "Категорія", "enabled": True, "template": "{{ category }}"},
    {"field": "Бренд", "enabled": True, "template": "{{ brand }}"},
    {"field": "Модель", "enabled": True, "template": "{{ model }}"},
    {"field": "Тип_плівки", "enabled": True, "template": "{{ film_type }}"},
    {"field": "Назва_позиції", "enabled": True, "template": "{{ title }}"},
    {"field": "Назва_позиції_укр", "enabled": False, "template": "{{ title }}"},
    {"field": "Пошукові_запити", "enabled": True, "template": "{{ tags }}"},
    {"field": "Пошукові_запити_укр", "enabled": False, "template": "{{ tags }}"},
    {"field": "Опис", "enabled": True, "template": "{{ description }}"},
    {"field": "Опис_укр", "enabled": False, "template": "{{ description }}"},
    {"field": "Код_товару", "enabled": False, "template": "{{ spec('Код_товару') }}"},
    {"field": "Тип_товару", "enabled": False, "template": "{{ film_type }}"},
    {"field": "Ціна", "enabled": False, "template": ""},
    {"field": "Валюта", "enabled": False, "template": ""},
    {"field": "Одиниця_виміру", "enabled": False, "template": ""},
    {"field": "Мінімальний_обсяг_замовлення", "enabled": False, "template": ""},
    {"field": "Оптова_ціна", "enabled": False, "template": ""},
    {"field": "Мінімальне_замовлення_опт", "enabled": False, "template": ""},
    {"field": "Посилання_зображення", "enabled": False, "template": ""},
    {"field": "Наявність", "enabled": False, "template": ""},
    {"field": "Кількість", "enabled": False, "template": ""},
    {"field": "Номер_групи", "enabled": False, "template": "{{ category_id }}"},
    {"field": "Назва_групи", "enabled": False, "template": "{{ category }}"},
    {"field": "Посилання_підрозділу", "enabled": False, "template": ""},
    {"field": "Можливість_поставки", "enabled": False, "template": ""},
    {"field": "Термін_поставки", "enabled": False, "template": ""},
    {"field": "Спосіб_пакування", "enabled": False, "template": ""},
    {"field": "Спосіб_пакування_укр", "enabled": False, "template": ""},
    {"field": "Унікальний_ідентифікатор", "enabled": False, "template": ""},
    {"field": "Ідентифікатор_товару", "enabled": False, "template": "{{ model_id }}"},
    {"field": "Ідентифікатор_підрозділу", "enabled": False, "template": ""},
    {"field": "Ідентифікатор_групи", "enabled": False, "template": "{{ category_id }}"},
    {"field": "Виробник", "enabled": False, "template": "{{ brand }}"},
    {"field": "Країна_виробник", "enabled": False, "template": ""},
    {"field": "Знижка", "enabled": False, "template": ""},
    {"field": "ID_групи_різновидів", "enabled": False, "template": ""},
    {"field": "Особисті_нотатки", "enabled": False, "template": ""},
    {"field": "Продукт_на_сайті", "enabled": False, "template": ""},
    {"field": "Термін_дії_знижки_від", "enabled": False, "template": ""},
    {"field": "Термін_дії_знижки_до", "enabled": False, "template": ""},
    {"field": "Ціна_від", "enabled": False, "template": ""},
    {"field": "Ярлик", "enabled": False, "template": ""},
    {"field": "HTML_заголовок", "enabled": False, "template": "{{ title }}"},
    {"field": "HTML_заголовок_укр", "enabled": False, "template": "{{ title }}"},
    {"field": "HTML_опис", "enabled": False, "template": "{{ description }}"},
    {"field": "HTML_опис_укр", "enabled": False, "template": "{{ description }}"},
    {"field": "Код_маркування_(GTIN)", "enabled": False, "template": ""},
    {"field": "Номер_пристрою_(MPN)", "enabled": False, "template": ""},
    {"field": "Вага,кг", "enabled": False, "template": "{{ spec('Вага, кг') }}"},
    {"field": "Ширина,см", "enabled": False, "template": "{{ spec('Ширина, см') }}"},
    {"field": "Висота,см", "enabled": False, "template": "{{ spec('Висота, см') }}"},
    {"field": "Довжина,см", "enabled": False, "template": "{{ spec('Довжина, см') }}"},
    {"field": "Де_знаходиться_товар", "enabled": False, "template": ""},
    {"field": "Назва_Характеристики", "enabled": False, "template": "{{ spec_items | map(attribute=0) | join('; ') }}"},
    {"field": "Одиниця_виміру,_Характеристики", "enabled": False, "template": ""},
    {"field": "Значення_Характеристики", "enabled": False, "template": "{{ spec_items | map(attribute=1) | join('; ') }}"},
]


def _copy_default_export_fields():
    defaults = []
    for item in DEFAULT_EXPORT_FIELDS:
        if not isinstance(item, dict):
            continue
        name = item.get("field")
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        template = item.get("template", "")
        if template is None:
            template = ""
        template = str(template)
        enabled = bool(item.get("enabled", False))
        languages = _normalize_export_field_languages(name, item.get("languages", _MISSING))
        defaults.append(
            {
                "field": name,
                "template": template,
                "enabled": enabled,
                "languages": languages,
            }
        )
    return defaults


EXCEL_FORMAT_LABEL = "Excel (.xlsx)"
CSV_FORMAT_LABEL = "CSV (.csv)"
JSON_FORMAT_LABEL = "JSON (.json)"
EXPORT_FORMAT_OPTIONS = (EXCEL_FORMAT_LABEL, CSV_FORMAT_LABEL, JSON_FORMAT_LABEL)


def get_available_export_formats():
    if not OPENPYXL_AVAILABLE or EXCEL_EXPORT_BLOCKED_MESSAGE:
        return [fmt for fmt in EXPORT_FORMAT_OPTIONS if fmt != EXCEL_FORMAT_LABEL]
    return list(EXPORT_FORMAT_OPTIONS)


def _title_tags_block(title: str, tags: str) -> dict:
    return {
        "title_template": _normalize_template_language_entry(title, fallback_value=title),
        "tags_template": _normalize_template_language_entry(tags, fallback_value=tags),
    }


def _build_title_tags_defaults(film_type_names, base_title, base_tags):
    base_block = _title_tags_block(base_title, base_tags)
    data = {
        "default": base_block.copy(),
        "by_film": {name: base_block.copy() for name in film_type_names},
        "by_category": {},
    }
    return data


def load_templates():
    path = _config_path(TEMPLATES_FILENAME)
    data = _load_json_config(path, lambda: deepcopy(DEFAULT_TEMPLATES), dict)

    for k, v in DEFAULT_TEMPLATES.items():
        if k not in data:
            data[k] = deepcopy(v)
    data["template_languages"] = _normalize_language_definitions(
        data.get("template_languages")
    )
    return data


def save_templates(dct):
    path = _config_path(TEMPLATES_FILENAME)
    _write_json_config(path, dct)


def _normalize_title_tags_block(block: dict, fallback: dict) -> dict:
    if not isinstance(block, dict):
        block = {}
    normalized = {}
    for key in ("title_template", "tags_template"):
        fallback_entry = fallback.get(key, {}) if isinstance(fallback, dict) else {}
        fallback_entry = _normalize_template_language_entry(fallback_entry)
        normalized_entry = _normalize_template_language_entry(
            block.get(key), fallback_value=fallback_entry.get("default", "")
        )
        fallback_languages = fallback_entry.get("languages", {})
        if isinstance(fallback_languages, dict):
            normalized_languages = normalized_entry.setdefault("languages", {})
            for code, text in fallback_languages.items():
                normalized_languages.setdefault(code, text)
        normalized[key] = normalized_entry
    return normalized


def load_title_tags_templates(templates: dict):
    film_type_names = [item.get("name") for item in templates.get("film_types", []) if item.get("name")]
    base_title = templates.get("title_template", DEFAULT_TEMPLATES["title_template"])
    base_tags = templates.get("tags_template", DEFAULT_TEMPLATES["tags_template"])
    defaults = _build_title_tags_defaults(film_type_names, base_title, base_tags)

    path = _config_path(TITLE_TAGS_FILENAME)
    data = _load_json_config(path, lambda: deepcopy(defaults), dict)

    changed = False

    legacy_film_blocks = {}
    for key in list(data.keys()):
        if key in ("default", "by_film", "by_category"):
            continue
        value = data.pop(key)
        if isinstance(value, dict):
            legacy_film_blocks[key] = value
            changed = True

    default_block = _normalize_title_tags_block(data.get("default"), defaults["default"])
    by_film_raw = data.get("by_film")
    if not isinstance(by_film_raw, dict):
        by_film_raw = {}
        changed = True
    by_category_raw = data.get("by_category")
    if not isinstance(by_category_raw, dict):
        by_category_raw = {}
        changed = True

    for name, block in legacy_film_blocks.items():
        by_film_raw[name] = _normalize_title_tags_block(block, default_block)

    normalized_by_film = {}
    for name, block in by_film_raw.items():
        normalized_by_film[name] = _normalize_title_tags_block(block, default_block)

    for name in film_type_names:
        if name not in normalized_by_film:
            normalized_by_film[name] = _normalize_title_tags_block({}, default_block)
            changed = True

    normalized_by_category = {}
    for cat_name, cat_block in by_category_raw.items():
        if not isinstance(cat_block, dict):
            changed = True
            continue
        cat_default = _normalize_title_tags_block(cat_block.get("default"), default_block)
        cat_films_raw = cat_block.get("by_film")
        if not isinstance(cat_films_raw, dict):
            cat_films_raw = {}
            changed = True
        normalized_cat_films = {}
        for film_name, film_block in cat_films_raw.items():
            normalized_cat_films[film_name] = _normalize_title_tags_block(
                film_block, cat_default
            )
        normalized_by_category[cat_name] = {
            "default": cat_default,
            "by_film": normalized_cat_films,
        }

    normalized = {
        "default": default_block,
        "by_film": normalized_by_film,
        "by_category": normalized_by_category,
    }

    if changed:
        save_title_tags_templates(normalized)

    return normalized


def save_title_tags_templates(dct):
    path = _config_path(TITLE_TAGS_FILENAME)
    _write_json_config(path, dct)


def resolve_title_tags(
    title_tags_templates: dict,
    templates: dict,
    category: Optional[str],
    film_type: str,
    language_codes: Optional[Iterable[str]] = None,
):
    fallback_title = templates.get("title_template", DEFAULT_TEMPLATES["title_template"])
    fallback_tags = templates.get("tags_template", DEFAULT_TEMPLATES["tags_template"])

    default_block = title_tags_templates.get("default", {})
    by_film = title_tags_templates.get("by_film", {})
    by_category = title_tags_templates.get("by_category", {})

    if not isinstance(default_block, dict):
        default_block = {}
    if not isinstance(by_film, dict):
        by_film = {}
    if not isinstance(by_category, dict):
        by_category = {}

    cat_block = {}
    if category:
        cat_block = by_category.get(category, {})
        if not isinstance(cat_block, dict):
            cat_block = {}
    cat_default = cat_block.get("default", {}) if isinstance(cat_block, dict) else {}
    if not isinstance(cat_default, dict):
        cat_default = {}
    cat_by_film = cat_block.get("by_film", {}) if isinstance(cat_block, dict) else {}
    if not isinstance(cat_by_film, dict):
        cat_by_film = {}
    cat_film_block = cat_by_film.get(film_type, {})
    if not isinstance(cat_film_block, dict):
        cat_film_block = {}

    film_block = by_film.get(film_type, {})
    if not isinstance(film_block, dict):
        film_block = {}

    codes = []
    if language_codes is not None:
        for code in language_codes:
            if not isinstance(code, str):
                continue
            stripped = code.strip()
            if not stripped:
                continue
            if stripped in codes:
                continue
            codes.append(stripped)
    if None not in codes:
        codes.append(None)

    def resolve_map(key: str, default_value: str) -> dict:
        resolved = {}
        for code in codes:
            value = None
            for block in (cat_film_block, cat_default, film_block, default_block):
                if not isinstance(block, dict):
                    continue
                candidate = _get_language_template_value(block.get(key), code, fallback_value=None)
                if isinstance(candidate, str):
                    value = candidate
                    break
            if value is None:
                value = default_value
            resolved[code] = value
        return resolved

    title_map = resolve_map("title_template", fallback_title)
    tags_map = resolve_map("tags_template", fallback_tags)
    return title_map, tags_map


def load_export_fields():
    path = _config_path(EXPORT_FIELDS_FILENAME)
    data = _load_json_config(path, _copy_default_export_fields, list)

    normalized = []
    changed = False
    for item in data:
        if not isinstance(item, dict):
            changed = True
            continue
        field_name = item.get("field") or item.get("name") or item.get("key")
        if field_name is None:
            changed = True
            continue
        field_name = str(field_name).strip()
        if not field_name:
            changed = True
            continue
        template = item.get("template", "")
        if template is None:
            template = ""
        template = str(template)
        enabled = bool(item.get("enabled", False))
        raw_languages = item.get("languages", _MISSING)
        languages = _normalize_export_field_languages(field_name, raw_languages)
        normalized.append(
            {
                "field": field_name,
                "template": template,
                "enabled": enabled,
                "languages": languages,
            }
        )

        sanitized_raw_languages = []
        if raw_languages is _MISSING:
            sanitized_raw_languages = None
        elif isinstance(raw_languages, str):
            raw_code = raw_languages.strip()
            sanitized_raw_languages = [raw_code] if raw_code else []
        elif isinstance(raw_languages, (list, tuple, set)):
            seen_codes = set()
            sanitized_raw_languages = []
            for lang in raw_languages:
                if not isinstance(lang, str):
                    continue
                code = lang.strip()
                if not code or code in seen_codes:
                    continue
                sanitized_raw_languages.append(code)
                seen_codes.add(code)
        else:
            sanitized_raw_languages = []
        if (
            field_name != item.get("field")
            or template != (item.get("template", "") or "")
            or enabled != bool(item.get("enabled", False))
            or (raw_languages is _MISSING and languages)
            or (sanitized_raw_languages is not None and sanitized_raw_languages != languages)
        ):
            changed = True

    if not normalized:
        normalized = _copy_default_export_fields()
        save_export_fields(normalized)
        return normalized

    if changed:
        save_export_fields(normalized)

    return normalized


def save_export_fields(fields: list):
    sanitized = []
    for item in fields:
        if not isinstance(item, dict):
            continue
        field_name = item.get("field") or item.get("name") or item.get("key")
        if field_name is None:
            continue
        field_name = str(field_name).strip()
        if not field_name:
            continue
        template = item.get("template", "")
        if template is None:
            template = ""
        template = str(template)
        enabled = bool(item.get("enabled", False))
        raw_languages = item.get("languages", _MISSING)
        languages = _normalize_export_field_languages(field_name, raw_languages)
        sanitized.append(
            {
                "field": field_name,
                "template": template,
                "enabled": enabled,
                "languages": languages,
            }
        )

    path = _config_path(EXPORT_FIELDS_FILENAME)
    _write_json_config(path, sanitized)

    return sanitized


# ============================ FORMULAS & GENERATION =================================

_FORMULA_PREFIX_RE = re.compile(r"^\s*=")
_IDENTIFIER_SANITIZE_RE = re.compile(r"[^\w]+", re.UNICODE)
_ASCII_SANITIZE_RE = re.compile(r"[^0-9a-z]+")

_TRANSLIT_TABLE = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "h",
    "ґ": "g",
    "д": "d",
    "е": "e",
    "є": "ie",
    "ж": "zh",
    "з": "z",
    "и": "y",
    "і": "i",
    "ї": "yi",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "kh",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "shch",
    "ь": "",
    "ю": "yu",
    "я": "ya",
    "ъ": "",
    "ы": "y",
    "э": "e",
    "ё": "yo",
}

_COMMON_SPEC_ALIAS = {
    "color": "color",
    "colour": "color",
    "kolir": "color",
    "колір": "color",
    "цвет": "color",
    "brand": "brand",
    "бренд": "brand",
    "weight": "weight",
    "вага": "weight",
    "вес": "weight",
    "material": "material",
    "матеріал": "material",
    "материал": "material",
    "thickness": "thickness",
    "товщина": "thickness",
    "толщина": "thickness",
    "sku": "sku",
    "код": "sku",
    "код_товару": "sku",
}


def _looks_like_formula(text: str) -> bool:
    return bool(text) and bool(_FORMULA_PREFIX_RE.match(text))


def _normalize_identifier(value: str) -> str:
    lowered = value.strip().lower()
    cleaned = _IDENTIFIER_SANITIZE_RE.sub("_", lowered)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned


def _transliterate_ascii(value: str) -> str:
    result = []
    for ch in value.lower():
        if ch.isdigit():
            result.append(ch)
            continue
        if "a" <= ch <= "z":
            result.append(ch)
            continue
        if ch in _TRANSLIT_TABLE:
            result.append(_TRANSLIT_TABLE[ch])
            continue
        if ch == "_" or ch.isspace():
            result.append("_")
        else:
            result.append("")
    ascii_candidate = "".join(result)
    ascii_candidate = _ASCII_SANITIZE_RE.sub("_", ascii_candidate)
    ascii_candidate = re.sub(r"_+", "_", ascii_candidate).strip("_")
    return ascii_candidate


def _build_formula_context(base_context):
    formula_context: Dict[str, object] = {}
    for key, value in base_context.items():
        if isinstance(value, _CallableDateTime):
            formula_context[key] = value()
            continue
        if callable(value):
            continue
        formula_context[key] = value

    if "brand" in base_context and "attr_brand" not in formula_context:
        formula_context["attr_brand"] = base_context.get("brand")
    if "model" in base_context and "attr_model" not in formula_context:
        formula_context["attr_model"] = base_context.get("model")

    specs = base_context.get("specs")
    if isinstance(specs, dict):
        for spec_key, spec_value in specs.items():
            if spec_value is None:
                continue
            key_str = str(spec_key)
            normalized = _normalize_identifier(key_str)
            ascii_name = _transliterate_ascii(key_str)
            for candidate in (normalized, ascii_name):
                if candidate:
                    formula_context.setdefault(f"attr_{candidate}", spec_value)
            alias_source = None
            if ascii_name and ascii_name in _COMMON_SPEC_ALIAS:
                alias_source = _COMMON_SPEC_ALIAS[ascii_name]
            elif normalized and normalized in _COMMON_SPEC_ALIAS:
                alias_source = _COMMON_SPEC_ALIAS[normalized]
            if alias_source:
                formula_context.setdefault(f"attr_{alias_source}", spec_value)

    category_name = base_context.get("category")
    if isinstance(category_name, str):
        cat_slug = _transliterate_ascii(category_name) or _normalize_identifier(category_name)
        if cat_slug:
            formula_context.setdefault(f"category_{cat_slug}", category_name)

    film_type = base_context.get("film_type")
    if isinstance(film_type, str):
        ft_slug = _transliterate_ascii(film_type) or _normalize_identifier(film_type)
        if ft_slug:
            formula_context.setdefault(f"film_type_{ft_slug}", film_type)

    return formula_context


def generate_export_rows(
    film_types: list,
    templates: dict,
    title_tags_templates: dict,
    export_fields: list,
    category_ids=None,
    brand_ids=None,
    model_ids=None,
    languages=None,
    progress_callback=None,
):
    film_types = list(film_types)
    pairs = collect_models(category_ids=category_ids, brand_ids=brand_ids, model_ids=model_ids)
    if not pairs:
        return [], []

    selected_language_codes = []
    selected_language_set = set()
    if languages is not None:
        raw_languages = languages
        if isinstance(raw_languages, str):
            raw_languages = [raw_languages]
        if isinstance(raw_languages, (list, tuple, set)):
            seen_langs = set()
            for lang in raw_languages:
                if not isinstance(lang, str):
                    continue
                code = lang.strip()
                if not code or code in seen_langs:
                    continue
                selected_language_codes.append(code)
                seen_langs.add(code)
    selected_language_set = set(selected_language_codes)

    enabled_fields = []
    for field in export_fields:
        if not isinstance(field, dict):
            continue
        name = field.get("field") or field.get("name") or field.get("key")
        if name is None:
            continue
        name = str(name).strip()
        if not name or not field.get("enabled"):
            continue
        template_str = field.get("template", "")
        if template_str is None:
            template_str = ""
        field_languages_raw = field.get("languages", _MISSING)
        field_languages = _normalize_export_field_languages(name, field_languages_raw)
        language_tuple = tuple(field_languages)
        if selected_language_set and language_tuple:
            if selected_language_set.isdisjoint(language_tuple):
                continue
        enabled_fields.append({"field": name, "template": str(template_str), "languages": language_tuple})

    if not enabled_fields:
        raise ValueError("Увімкніть хоча б одне поле експорту.")

    specs_map = load_specs_map([mid for _brand, _model, _cat, mid, _bid, _cid in pairs])
    title_tags_cache = {}
    desc_template_cache = {}
    field_template_cache = {}
    rows = []
    now_value = datetime.now()
    now_for_context = _CallableDateTime(now_value)

    column_order = [field["field"] for field in enabled_fields]

    descriptions = templates.get("descriptions", {})
    global_desc_block = {}
    if isinstance(descriptions, dict):
        candidate = descriptions.get(GLOBAL_DESCRIPTION_KEY)
        if isinstance(candidate, dict):
            global_desc_block = candidate
    template_languages = _normalize_language_definitions(templates.get("template_languages"))
    template_language_codes = [item.get("code") for item in template_languages if item.get("code")]
    template_language_codes = [code for code in template_language_codes if isinstance(code, str) and code.strip()]
    template_language_codes = [code.strip() for code in template_language_codes]
    language_iteration = list(template_language_codes)
    if None not in language_iteration:
        language_iteration.append(None)
    primary_language = template_language_codes[0] if template_language_codes else None

    def _language_label(code: Optional[str]) -> str:
        if code is None or code == "":
            return "за замовчуванням"
        return code

    def _language_key_suffix(code: Optional[str]) -> str:
        if code is None or code == "":
            return "default"
        return re.sub(r"\W+", "_", code)

    def _value_for_language(values: dict) -> str:
        if primary_language is not None:
            primary_value = values.get(primary_language)
            if isinstance(primary_value, str):
                return primary_value
        default_value = values.get(None)
        if isinstance(default_value, str):
            return default_value
        for value in values.values():
            if isinstance(value, str):
                return value
        return ""

    total_steps = len(pairs) * len(film_types)
    progress_count = 0
    if progress_callback is not None:
        progress_callback(progress_count, total_steps)

    for brand, model, cat, mid, brand_id, cat_id in pairs:
        specs = specs_map.get(mid, {})
        if not isinstance(specs, dict):
            specs = {}
        spec_items = list(specs.items())

        def spec_lookup(key, default=""):
            return specs.get(key, default)

        cat_desc_block: Dict[str, dict] = {}
        if isinstance(descriptions, dict):
            cat_block = descriptions.get(cat)
            if isinstance(cat_block, dict):
                cat_desc_block = cat_block

        def _entry_from_blocks(key: str):
            sources = [cat_desc_block, global_desc_block]
            for block in sources:
                if isinstance(block, dict):
                    value = block.get(key)
                    if isinstance(value, (dict, str)):
                        return value
            return None
        for f in film_types:
            film_type = f if isinstance(f, str) else str(f)
            cache_key = (cat, film_type)
            if cache_key not in title_tags_cache:
                title_map, tags_map = resolve_title_tags(
                    title_tags_templates,
                    templates,
                    cat,
                    film_type,
                    template_language_codes,
                )
                compiled_map = {}
                map_keys = set(title_map.keys()) | set(tags_map.keys())
                for code in map_keys:
                    title_tpl_str = title_map.get(code)
                    tags_tpl_str = tags_map.get(code)
                    try:
                        title_tpl = Template(title_tpl_str or "")
                    except TemplateError as exc:
                        label = _language_label(code)
                        raise ValueError(
                            f"Помилка в шаблоні заголовку для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                        ) from exc
                    try:
                        tags_tpl = Template(tags_tpl_str or "")
                    except TemplateError as exc:
                        label = _language_label(code)
                        raise ValueError(
                            f"Помилка в шаблоні тегів для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                        ) from exc
                    compiled_map[code] = (title_tpl, tags_tpl)
                title_tags_cache[cache_key] = compiled_map
            compiled_map = title_tags_cache[cache_key]

            render_kwargs = dict(film_type=film_type, brand=brand, model=model, category=cat)
            title_values = {}
            tags_values = {}
            for code in language_iteration:
                tpl_pair = compiled_map.get(code) or compiled_map.get(None)
                if tpl_pair is None and compiled_map:
                    tpl_pair = next(iter(compiled_map.values()))
                if tpl_pair is None:
                    raise ValueError(
                        f"Не знайдено шаблон заголовку/тегів для категорії \"{cat}\" і типу \"{film_type}\"."
                    )
                title_tpl, tags_tpl = tpl_pair
                try:
                    title_rendered = title_tpl.render(**render_kwargs)
                except TemplateError as exc:
                    label = _language_label(code)
                    raise ValueError(
                        f"Не вдалося згенерувати заголовок для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                    ) from exc
                try:
                    tags_rendered = tags_tpl.render(**render_kwargs)
                except TemplateError as exc:
                    label = _language_label(code)
                    raise ValueError(
                        f"Не вдалося згенерувати теги для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                    ) from exc
                title_values[code] = title_rendered
                tags_values[code] = tags_rendered

            desc_key = (cat, film_type)
            desc_compiled = desc_template_cache.get(desc_key)
            if desc_compiled is None:
                film_entry = _entry_from_blocks(film_type)
                default_entry = _entry_from_blocks("default")
                fallback_desc = "Плівка для {{ brand }} {{ model }}"
                desc_compiled = {}
                for code in language_iteration:
                    template_str = None
                    for entry in (film_entry, default_entry):
                        if entry is None:
                            continue
                        candidate = _get_language_template_value(entry, code, fallback_value=None)
                        if isinstance(candidate, str):
                            template_str = candidate
                            break
                    if template_str is None:
                        template_str = fallback_desc
                    try:
                        desc_compiled[code] = Template(template_str or "")
                    except TemplateError as exc:
                        label = _language_label(code)
                        raise ValueError(
                            f"Помилка в шаблоні опису для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                        ) from exc
                desc_template_cache[desc_key] = desc_compiled
            desc_values = {}
            for code in language_iteration:
                tpl = desc_compiled.get(code) or desc_compiled.get(None)
                if tpl is None and desc_compiled:
                    tpl = next(iter(desc_compiled.values()))
                if tpl is None:
                    raise ValueError(
                        f"Не знайдено шаблон опису для категорії \"{cat}\" і типу \"{film_type}\"."
                    )
                try:
                    desc_rendered = tpl.render(film_type=film_type, brand=brand, model=model, category=cat)
                except TemplateError as exc:
                    label = _language_label(code)
                    raise ValueError(
                        f"Не вдалося сформувати опис для категорії \"{cat}\", типу \"{film_type}\" і мови \"{label}\": {exc}"
                    ) from exc
                desc_values[code] = desc_rendered

            default_title = _value_for_language(title_values)
            default_tags = _value_for_language(tags_values)
            default_desc = _value_for_language(desc_values)

            context = {
                "brand": brand,
                "brand_id": brand_id,
                "model": model,
                "model_id": mid,
                "category": cat,
                "category_id": cat_id,
                "film_type": film_type,
                "title": default_title,
                "description": default_desc,
                "tags": default_tags,
                "specs": specs,
                "spec_items": spec_items,
                "spec": spec_lookup,
                "row_number": len(rows) + 1,
                "now": now_for_context,
                "now_value": now_value,
                "language": None,
                "selected_languages": tuple(selected_language_codes),
                "titles_localized": dict(title_values),
                "descriptions_localized": dict(desc_values),
                "tags_localized": dict(tags_values),
                "available_template_languages": tuple(template_language_codes),
            }

            context["relativedelta"] = _relativedelta_helper

            for code, value in title_values.items():
                suffix = _language_key_suffix(code)
                context[f"title_{suffix}"] = value
            for code, value in desc_values.items():
                suffix = _language_key_suffix(code)
                context[f"description_{suffix}"] = value
            for code, value in tags_values.items():
                suffix = _language_key_suffix(code)
                context[f"tags_{suffix}"] = value

            row_values = []
            render_context_cache = {None: context}
            formula_context_cache = {}
            for field in enabled_fields:
                field_name = field["field"]
                tpl_str = field.get("template", "")
                field_languages = field.get("languages") or ()
                if isinstance(field_languages, str):
                    field_languages = (field_languages,)
                language_code = None
                if field_languages and len(field_languages) == 1:
                    raw_code = field_languages[0]
                    if isinstance(raw_code, str):
                        stripped_code = raw_code.strip()
                        language_code = stripped_code or None
                context_for_field = render_context_cache.get(language_code)
                if context_for_field is None:
                    context_for_field = dict(context)
                    context_for_field["language"] = language_code
                    effective_code = language_code
                    title_for_language = title_values.get(effective_code)
                    if title_for_language is None and effective_code is not None:
                        title_for_language = title_values.get(None)
                    if title_for_language is None:
                        title_for_language = default_title
                    desc_for_language = desc_values.get(effective_code)
                    if desc_for_language is None and effective_code is not None:
                        desc_for_language = desc_values.get(None)
                    if desc_for_language is None:
                        desc_for_language = default_desc
                    tags_for_language = tags_values.get(effective_code)
                    if tags_for_language is None and effective_code is not None:
                        tags_for_language = tags_values.get(None)
                    if tags_for_language is None:
                        tags_for_language = default_tags
                    context_for_field["title"] = title_for_language
                    context_for_field["description"] = desc_for_language
                    context_for_field["tags"] = tags_for_language
                    render_context_cache[language_code] = context_for_field
                if tpl_str:
                    if _looks_like_formula(tpl_str):
                        formula_context = formula_context_cache.get(language_code)
                        if formula_context is None:
                            formula_context = _build_formula_context(context_for_field)
                            formula_context_cache[language_code] = formula_context
                        try:
                            value = FormulaEngine.evaluate(tpl_str, formula_context)
                        except FormulaError as exc:
                            raise ValueError(
                                f"Помилка у формулі поля \"{field_name}\": {exc}"
                            ) from exc
                    else:
                        tpl = field_template_cache.get(tpl_str)
                        if tpl is None:
                            try:
                                tpl = Template(tpl_str)
                            except TemplateError as exc:
                                raise ValueError(
                                    f"Помилка в шаблоні поля \"{field_name}\": {exc}"
                                ) from exc
                            field_template_cache[tpl_str] = tpl
                        try:
                            value = tpl.render(**context_for_field)
                        except TemplateError as exc:
                            raise ValueError(
                                f"Не вдалося згенерувати значення поля \"{field_name}\": {exc}"
                            ) from exc
                else:
                    value = context_for_field.get(field_name)
                    if value is None:
                        value = context.get(field_name, "")
                if value is None:
                    value = ""
                elif not isinstance(value, str):
                    value = str(value)
                row_values.append(value)
            rows.append(row_values)
            progress_count += 1
            if progress_callback is not None:
                progress_callback(progress_count, total_steps)

    return rows, column_order


def _row_to_values(record, columns):
    if isinstance(record, dict):
        return [record.get(name, "") for name in columns]
    if isinstance(record, (list, tuple)):
        values = list(record)
        if len(values) < len(columns):
            values.extend([""] * (len(columns) - len(values)))
        elif len(values) > len(columns):
            values = values[: len(columns)]
        return values
    return ["" for _ in columns]


def _make_unique_column_keys(columns):
    counts = {}
    unique_keys = []
    for name in columns:
        count = counts.get(name, 0) + 1
        counts[name] = count
        if count == 1:
            unique_keys.append(name)
        else:
            unique_keys.append(f"{name}__{count}")
    return unique_keys


def ensure_folder(path: str):
    if path and not os.path.exists(path):
        try:
            os.makedirs(path, exist_ok=True)
        except OSError as exc:
            message = f"Не вдалося підготувати теку для експорту: {exc}"
            raise ExportError(EXPORT_ERR_FOLDER_PREP, message) from exc


def export_products(records: list, columns: list, fmt: str, folder: str):
    ensure_folder(folder)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.join(folder, f"products_{ts}")

    if fmt == EXCEL_FORMAT_LABEL:
        if not OPENPYXL_AVAILABLE or EXCEL_EXPORT_BLOCKED_MESSAGE:
            message = EXCEL_EXPORT_BLOCKED_MESSAGE or "Експорт у Excel недоступний."
            detail = OPENPYXL_IMPORT_ERROR_DETAIL
            if detail and detail not in message:
                message = f"{message} (деталі: {detail})"
            message = f"{message}\n{OPENPYXL_INSTALL_HINT}{CSV_JSON_FALLBACK_NOTE}"
            raise ExportError(EXPORT_ERR_NO_OPENPYXL, message)

        out_products = base + ".xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        column_widths: List[int] = []
        if columns:
            sheet.append(columns)
            column_widths = [max(10, min(60, len(str(col) if col is not None else ""))) for col in columns]

        for record in records:
            row = _row_to_values(record, columns)
            if column_widths:
                for idx, value in enumerate(row):
                    length = len(str(value) if value is not None else "")
                    column_widths[idx] = min(60, max(column_widths[idx], max(10, length)))
            sheet.append(row)

        if columns:
            try:
                sheet.freeze_panes = "A2"
            except Exception:
                LOGGER.debug("Не вдалося зафіксувати рядок заголовків у Excel", exc_info=True)

            try:
                from openpyxl.utils import get_column_letter as _get_column_letter
            except Exception:  # pragma: no cover - fallback when utils unavailable
                _get_column_letter = None
            try:
                from openpyxl.styles import Alignment as _Alignment
            except Exception:  # pragma: no cover - fallback when styles unavailable
                _Alignment = None

            def _column_letter(idx: int) -> str:
                if _get_column_letter is not None:
                    return _get_column_letter(idx)
                base = ord("A") + (idx - 1)
                if 0 <= base < 26 + ord("A"):
                    return chr(base)
                return f"COL{idx}"

            try:
                sheet.auto_filter.ref = f"A1:{_column_letter(len(columns))}1"
            except Exception:
                LOGGER.debug("Не вдалося застосувати автофільтр до Excel-аркуша", exc_info=True)

            if _Alignment is not None:
                alignment = _Alignment(wrap_text=True)
            else:
                class _SimpleAlignment:  # pragma: no cover - fallback for limited stubs
                    def __init__(self, wrap_text: bool = False):
                        self.wrap_text = wrap_text

                alignment = _SimpleAlignment(wrap_text=True)

            if hasattr(sheet, "column_dimensions") and column_widths:
                for idx, width in enumerate(column_widths, start=1):
                    letter = _column_letter(idx)
                    dimension = None
                    try:
                        dimension = sheet.column_dimensions.get(letter)
                    except Exception:
                        dimension = None
                    if dimension is None:
                        try:
                            dimension = sheet.column_dimensions[letter]
                        except Exception:
                            LOGGER.debug("Не вдалося призначити ширину колонки %s", letter, exc_info=True)
                            continue
                    try:
                        dimension.width = width
                    except Exception:
                        try:
                            sheet.column_dimensions[letter].width = width
                        except Exception:
                            LOGGER.debug("Не вдалося призначити ширину колонки %s", letter, exc_info=True)

            if alignment is not None and hasattr(sheet, "iter_rows"):
                try:
                    max_row = getattr(sheet, "max_row", len(getattr(sheet, "rows", [])))
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=len(columns)):
                        for cell in row:
                            try:
                                cell.alignment = alignment
                            except Exception:
                                pass
                except Exception:
                    LOGGER.debug("Не вдалося застосувати перенесення тексту у клітинках Excel", exc_info=True)

        try:
            workbook.save(out_products)
        except PermissionError as exc:
            message = (
                "Не вдалося зберегти Excel-файл: доступ заборонено. Закрийте файл, якщо він відкритий, та спробуйте знову."
            )
            raise ExportError(EXPORT_ERR_PERMISSION, message) from exc
        except OSError as exc:  # pragma: no cover - defensive
            message = f"Не вдалося зберегти Excel-файл: {exc}"
            raise ExportError(EXPORT_ERR_OS_ERROR, message) from exc
        finally:
            try:
                workbook.close()
            except Exception:
                pass
    elif fmt == CSV_FORMAT_LABEL:
        out_products = base + ".csv"
        try:
            with open(out_products, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if columns:
                    writer.writerow(columns)
                for record in records:
                    row = _row_to_values(record, columns)
                    writer.writerow(row)
        except PermissionError as exc:
            message = (
                "Не вдалося зберегти CSV-файл: доступ заборонено. Закрийте файл, якщо він відкритий, та спробуйте знову."
            )
            raise ExportError(EXPORT_ERR_PERMISSION, message) from exc
        except OSError as exc:
            message = f"Не вдалося зберегти CSV-файл: {exc}"
            raise ExportError(EXPORT_ERR_OS_ERROR, message) from exc
    elif fmt == JSON_FORMAT_LABEL:
        out_products = base + ".json"
        json_records = []
        json_columns = _make_unique_column_keys(columns)
        for record in records:
            values = _row_to_values(record, columns)
            json_records.append({key: value for key, value in zip(json_columns, values)})
        try:
            with open(out_products, "w", encoding="utf-8") as f:
                json.dump(json_records, f, ensure_ascii=False, indent=2)
        except PermissionError as exc:
            message = (
                "Не вдалося зберегти JSON-файл: доступ заборонено. Закрийте файл, якщо він відкритий, та спробуйте знову."
            )
            raise ExportError(EXPORT_ERR_PERMISSION, message) from exc
        except OSError as exc:
            message = f"Не вдалося зберегти JSON-файл: {exc}"
            raise ExportError(EXPORT_ERR_OS_ERROR, message) from exc
    else:
        message = f"Невідомий формат експорту: {fmt}"
        raise ExportError(EXPORT_ERR_UNKNOWN_FORMAT, message)

    return out_products
